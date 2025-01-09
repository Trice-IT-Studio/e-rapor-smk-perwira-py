from openpyxl import Workbook
from openpyxl.styles import Alignment
import os


def laporan_tengah_toexcel(
    th_obj, sm_obj, mapel_obj, kel_mapel_obj, kelas_obj, siswa_obj, **kwargs
):
    wb = Workbook()
    ws = wb.active

    # keyword argumetns
    tipe_nilai = kwargs.get("tipe_nilai", None)

    th_value = "".join(th_obj.value.split("/"))

    if not os.path.isdir("data_nilai"):
        os.makedirs("data_nilai")

    # SAVE NILAI PTS TO EXCEL ###
    if tipe_nilai == 1:
        # PREPARE TITLE
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = "LAPORAN HASIL BELAJAR TENGAH SEMESTER"
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # SISWA INFO
        ws.merge_cells("A2:B2")
        nama_siswa = ws["A2"]
        nama_siswa.value = "Nama Peserta Didik"
        ws.merge_cells("C2:D2")
        nama_siswa = ws["C2"]
        nama_siswa.value = f": {siswa_obj.full_name}"

        ws.merge_cells("A3:B3")
        nis_siswa = ws["A3"]
        nis_siswa.value = "NIS / NISN"
        ws.merge_cells("C3:D3")
        nis_siswa = ws["C3"]
        nis_siswa.value = f": {siswa_obj.nis} / {siswa_obj.nisn}"

        ws.merge_cells("A4:B4")
        kel_sm = ws["A4"]
        kel_sm.value = "Kelas/Semester"
        ws.merge_cells("C4:D4")
        kel_sm = ws["C4"]
        try:
            kel_sm.value = f": {siswa_obj.kesetaraan.value} {siswa_obj.kelas_belajar.name} / {sm_obj.value}"
        except Exception:
            kel_sm.value = f": ? {siswa_obj.kelas_belajar.name} / {sm_obj.value}"

        # TABLE HEADERS
        ws.merge_cells("A5:A6")
        active_cell = ws["A5"]
        active_cell.value = "No"

        ws.merge_cells("B5:B6")
        active_cell = ws["B5"]
        active_cell.value = "Mata Pelajaran"

        ws.merge_cells("C5:E5")
        active_cell = ws["C5"]
        active_cell.value = "Pengetahuan"

        active_cell = ws["C6"]
        active_cell.value = "Jml. Kd"

        active_cell = ws["D6"]
        active_cell.value = "Nilai"

        active_cell = ws["E6"]
        active_cell.value = "Predikat"

        ws.merge_cells("F5:H5")
        active_cell = ws["F5"]
        active_cell.value = "Keterampilan"

        active_cell = ws["F6"]
        active_cell.value = "Jml. Kd"

        active_cell = ws["G6"]
        active_cell.value = "Nilai"

        active_cell = ws["H6"]
        active_cell.value = "Predikat"

        ws.merge_cells("I5:I6")
        active_cell = ws["I5"]
        active_cell.value = "Rata-Rata (N)"

        ccount = 0
        mapel_count = 1
        for index, kel_mapel in enumerate(kel_mapel_obj):
            # INDEX COL
            ws.merge_cells(f"A{7+ccount}:I{7+ccount}")
            active_cell = ws[f"A{7+ccount}"]
            active_cell.value = f"Kelompok {chr(65 + index)} ({kel_mapel.category})"

            # VALUE COL
            for mapel in kel_mapel.anggota_kelompok:
                if mapel.kelas_id == kelas_obj.id:
                    ccount += 1

                    active_cell = ws[f"A{7+ccount}"]
                    active_cell.value = mapel_count
                    active_cell = ws[f"B{7+ccount}"]
                    active_cell.value = mapel.name
                    active_cell = ws[f"C{7+ccount}"]
                    active_cell.value = len(mapel.kompetensi_dasar)
                    active_cell = ws[f"F{7+ccount}"]
                    active_cell.value = len(mapel.kompetensi_dasar)

                    for nlk in mapel.nilai_lingkup_materi:
                        if (
                            nlk.siswa_id == siswa_obj.id
                            and nlk.th_id == th_obj.id
                            and nlk.sm_id == sm_obj.id
                        ):
                            active_cell = ws[f"D{7+ccount}"]
                            if nlk.tipe_nilai == "p":
                                active_cell.value = nlk.value
                                active_cell = ws[f"E{7+ccount}"]
                                predikat = ""
                                if nlk.value > 80:
                                    predikat = "A"
                                elif nlk.value > 70 and nlk.value < 80:
                                    predikat = "B"
                                elif nlk.value > 60 and nlk.value < 70:
                                    predikat = "C"
                                else:
                                    predikat = "D"
                                active_cell.value = predikat
                            active_cell = ws[f"G{7+ccount}"]
                            if nlk.tipe_nilai == "k":
                                active_cell.value = nlk.value
                                active_cell = ws[f"H{7+ccount}"]
                                predikat = ""
                                if nlk.value > 80:
                                    predikat = "A"
                                elif nlk.value > 70 and nlk.value < 80:
                                    predikat = "B"
                                elif nlk.value > 60 and nlk.value < 70:
                                    predikat = "C"
                                else:
                                    predikat = "D"
                                active_cell.value = predikat
                    mapel_count += 1
            ccount += 1

        # EXPORT EXCEL
        wb.save(
            f"data_nilai/LaporanBelajarTS_{th_value}-{sm_obj.value}_{kelas_obj.name}_{siswa_obj.full_name}.xlsx"
        )
        return

    return Exception("No tipe nilai provided or not implemented")


def sumatif_tengah_toexcel(
    th_obj, sm_obj, mapel_obj, kel_mapel_obj, kelas_obj, siswa_obj, **kwargs
):
    """
    create a report of nilai siswa (Tengah semester & akhir semester)
    and save it as an excel formated file.

    th_id           tahun ajaran object
    sm_id           semester object
    mapel_obj       mata pelajaran object
    kelas_obj       kelas object
    siswa_obj       siswa object
    """
    wb = Workbook()
    ws = wb.active

    # keyword argumetns
    tipe_nilai = kwargs.get("tipe_nilai", None)

    th_value = "".join(th_obj.value.split("/"))

    if not os.path.isdir("data_nilai"):
        os.makedirs("data_nilai")

    # SAVE NILAI PTS TO EXCEL ###
    if tipe_nilai == 2:
        # PREPARE TITLE
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "LAPORAN HASIL BELAJAR\nSUMATIF TENGAH SEMESTER\n(RAPOR)"
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # STUDENTS INFORMATION
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 5
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["f"].width = 20

        active_cell = ws["B2"]
        active_cell.value = "Nama Peserta Didik"
        ws.merge_cells("C2:D2")
        active_cell = ws["C2"]
        active_cell.value = f": {siswa_obj.full_name}"

        active_cell = ws["B3"]
        active_cell.value = "NISN"
        ws.merge_cells("C3:D3")
        active_cell = ws["C3"]
        active_cell.value = f": {siswa_obj.nisn}"

        active_cell = ws["B4"]
        active_cell.value = "Sekolah"
        ws.merge_cells("C4:D4")
        active_cell = ws["C4"]
        active_cell.value = ": SMK Perwira Jakarta"

        active_cell = ws["B5"]
        active_cell.value = f"Alamat"
        ws.merge_cells("C5:D5")
        active_cell = ws["C5"]
        active_cell.value = f": {siswa_obj.alamat}"

        active_cell = ws["E2"]
        active_cell.value = "Kelas"
        active_cell = ws["F2"]
        active_cell.value = f": {kelas_obj.name}"

        active_cell = ws["E3"]
        active_cell.value = "Fase"
        active_cell = ws["F3"]
        active_cell.value = ": E"

        active_cell = ws["E4"]
        active_cell.value = "Semester"
        active_cell = ws["F4"]
        active_cell.value = f": {sm_obj.value}"

        active_cell = ws["E5"]
        active_cell.value = "Tahun Ajaran"
        active_cell = ws["F5"]
        active_cell.value = f": {th_obj.value}"

        # TABLE HEADS ###
        active_cell = ws["A6"]
        active_cell.value = "No"

        active_cell = ws["B6"]
        active_cell.value = "Muatan Pelajaran"

        active_cell = ws["C6"]
        ws.merge_cells("C6:D6")
        active_cell.value = "Nilai AKhir"

        active_cell = ws["E6"]
        ws.merge_cells("E6:F6")
        active_cell.value = "Capaian Kompetensi"

        # NILAI SECTION
        ccount = 0
        mapel_count = 1
        for index, kel_mapel in enumerate(kel_mapel_obj):
            # INDEX COL
            active_cell = ws[f"A{7+ccount}"]
            active_cell.value = chr(65 + index)

            # VALUE COL
            ws.merge_cells(f"B{7+ccount}:F{7+ccount}")
            active_cell = ws[f"B{7+ccount}"]
            active_cell.value = f"Kelompok Mata Pelajaran {kel_mapel.category}"
            for mapel in kel_mapel.anggota_kelompok:
                if mapel.kelas_id == kelas_obj.id:
                    ccount += 1

                    active_cell = ws[f"A{7+ccount}"]
                    active_cell.value = mapel_count
                    active_cell = ws[f"B{7+ccount}"]
                    active_cell.value = mapel.name

                    for nt in mapel.nilai_tengah:
                        if (
                            nt.siswa_id == siswa_obj.id
                            and nt.th_id == th_obj.id
                            and nt.sm_id == sm_obj.id
                        ):
                            ws.merge_cells(f"C{7+ccount}:D{7+ccount}")
                            active_cell = ws[f"C{7+ccount}"]
                            active_cell.value = nt.value

                    for ck in mapel.capaian_kompetensi:
                        if ck.th_id == th_obj.id and ck.sm_id == sm_obj.id:
                            ws.merge_cells(f"E{7+ccount}:F{7+ccount}")
                            active_cell = ws[f"E{7+ccount}"]
                            active_cell.value = ck.desc
                    mapel_count += 1
            ccount += 1

        # EXPORT EXCEL
        wb.save(
            f"data_nilai/PTS_{th_value}-{sm_obj.value}_{kelas_obj.name}_{siswa_obj.full_name}.xlsx"
        )
        return

    return Exception("No tipe nilai provided or not implemented")


def sumatif_akhir_toexcel(
    th_obj, sm_obj, mapel_obj, kel_mapel_obj, kelas_obj, siswa_obj, **kwargs
):
    """
    create a report of nilai siswa (akhir semester)
    and save it as an excel formated file.

    th_id           tahun ajaran object
    sm_id           semester object
    mapel_obj       mata pelajaran object
    kelas_obj       kelas object
    siswa_obj       siswa object
    """
    wb = Workbook()
    ws = wb.active

    # keyword argumetns
    tipe_nilai = kwargs.get("tipe_nilai", None)

    th_value = "".join(th_obj.value.split("/"))

    if not os.path.isdir("data_nilai"):
        os.makedirs("data_nilai")

    # SAVE NILAI PTS TO EXCEL ###
    if tipe_nilai == 3:
        # PREPARE TITLE
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "LAPORAN HASIL BELAJAR\nSUMATIF AKHIR SEMESTER\n(RAPOR)"
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # STUDENTS INFORMATION
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 5
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["f"].width = 20

        active_cell = ws["B2"]
        active_cell.value = "Nama Peserta Didik"
        ws.merge_cells("C2:D2")
        active_cell = ws["C2"]
        active_cell.value = f": {siswa_obj.full_name}"

        active_cell = ws["B3"]
        active_cell.value = "NISN"
        ws.merge_cells("C3:D3")
        active_cell = ws["C3"]
        active_cell.value = f": {siswa_obj.nisn}"

        active_cell = ws["B4"]
        active_cell.value = "Sekolah"
        ws.merge_cells("C4:D4")
        active_cell = ws["C4"]
        active_cell.value = ": SMK Perwira Jakarta"

        active_cell = ws["B5"]
        active_cell.value = f"Alamat"
        ws.merge_cells("C5:D5")
        active_cell = ws["C5"]
        active_cell.value = f": {siswa_obj.alamat}"

        active_cell = ws["E2"]
        active_cell.value = "Kelas"
        active_cell = ws["F2"]
        active_cell.value = f": {kelas_obj.name}"

        active_cell = ws["E3"]
        active_cell.value = "Fase"
        active_cell = ws["F3"]
        active_cell.value = ": E"

        active_cell = ws["E4"]
        active_cell.value = "Semester"
        active_cell = ws["F4"]
        active_cell.value = f": {sm_obj.value}"

        active_cell = ws["E5"]
        active_cell.value = "Tahun Ajaran"
        active_cell = ws["F5"]
        active_cell.value = f": {th_obj.value}"

        # TABLE HEADS ###
        active_cell = ws["A6"]
        active_cell.value = "No"

        active_cell = ws["B6"]
        active_cell.value = "Muatan Pelajaran"

        active_cell = ws["C6"]
        ws.merge_cells("C6:D6")
        active_cell.value = "Nilai AKhir"

        active_cell = ws["E6"]
        ws.merge_cells("E6:F6")
        active_cell.value = "Capaian Kompetensi"

        # NILAI SECTION
        ccount = 0
        mapel_count = 1
        for index, kel_mapel in enumerate(kel_mapel_obj):
            # INDEX COL
            active_cell = ws[f"A{7+ccount}"]
            active_cell.value = chr(65 + index)

            # VALUE COL
            ws.merge_cells(f"B{7+ccount}:F{7+ccount}")
            active_cell = ws[f"B{7+ccount}"]
            active_cell.value = f"Kelompok Mata Pelajaran {kel_mapel.category}"
            for mapel in kel_mapel.anggota_kelompok:
                if mapel.kelas_id == kelas_obj.id:
                    ccount += 1

                    active_cell = ws[f"A{7+ccount}"]
                    active_cell.value = mapel_count
                    active_cell = ws[f"B{7+ccount}"]
                    active_cell.value = mapel.name

                    for nt in mapel.nilai_akhir:
                        if (
                            nt.siswa_id == siswa_obj.id
                            and nt.th_id == th_obj.id
                            and nt.sm_id == sm_obj.id
                        ):
                            ws.merge_cells(f"C{7+ccount}:D{7+ccount}")
                            active_cell = ws[f"C{7+ccount}"]
                            active_cell.value = nt.value

                    for ck in mapel.capaian_kompetensi:
                        if ck.th_id == th_obj.id and ck.sm_id == sm_obj.id:
                            ws.merge_cells(f"E{7+ccount}:F{7+ccount}")
                            active_cell = ws[f"E{7+ccount}"]
                            active_cell.value = ck.desc
                    mapel_count += 1
            ccount += 1

        # EXPORT EXCEL
        wb.save(
            f"data_nilai/SAS_{th_value}-{sm_obj.value}_{kelas_obj.name}_{siswa_obj.full_name}.xlsx"
        )
        return

    return Exception("No tipe nilai provided or not implemented")
